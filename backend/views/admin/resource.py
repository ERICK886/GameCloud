import json

import openpyxl
from django.http import JsonResponse

from backend.func.config import get_config
from backend.func.token import check_admin, check_token
from backend.models import Resource, Image, Video, Tag, Category, Cover


def admin_resource_get(request):
    """
    管理员获取资源
    :param request:
    :return:
    """
    token = request.headers.get('Authorization')
    if not check_token(token):
        return JsonResponse({'code': 401, 'msg': 'token认证失败'})
    if not check_admin(token):
        return JsonResponse({'code': 401, 'msg': '权限不足'})
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        site_url = get_config('SITE_URL')
        if data.get('resourceId'):
            resourceId = data.get('resourceId')
        else:
            resourceId = ''

        if resourceId:
            resource = Resource.objects.filter(id=data.get('resourceId')).first()
            resource_list = {
                'id': resource.id,
                'name': resource.name,
                'description': resource.description,
                'active_code': resource.active_code,
                'aliUrl': resource.aliyun_url,
                'baiduUrl': resource.baidu_url,
                'tianyiUrl': resource.tianyi_url,
                'category': resource.category.id,
                'cover': [{
                    'id': resource.cover.id,
                    'url': site_url + '/uploads' + resource.cover.cover.url,
                    'name': resource.cover.name,
                    'status': 'finished'
                }],
                'images': [{
                    'id': image.id,
                    'url': site_url + '/uploads' + image.image.url,
                    'name': image.name,
                    'status': 'finished'
                } for image in resource.images.all()],
                'videos': [{
                    'id': video.id,
                    'url': site_url + '/uploads' + video.video.url,
                    'status': 'finished',
                    'name': video.name
                } for video in resource.videos.all()],
                'status': resource.status,
                'tag': [tag.id for tag in resource.tags.all()],
                'create_time': resource.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                'update_time': resource.update_time.strftime('%Y-%m-%d%H:%M:%S')
            }
            return JsonResponse({'code': 200, 'msg': '获取成功', 'data': resource_list})
        if data.get('kwd'):
            kwd = data.get('kwd')
        else:
            kwd = ''
        if data.get('order'):
            order = data.get('order')
        else:
            order = 'asc'
        if data.get('orderby'):
            orderby = data.get('orderby')
        else:
            orderby = 'id'

        if order == 'asc':
            order = ''
        else:
            order = '-'

        resource_list = []
        resources = Resource.objects.filter(name__icontains=kwd).order_by(order + orderby)
        for resource in resources:
            resource_list.append({
                'id': resource.id,
                'name': resource.name,
                'description': resource.description,
                'content': resource.content,
                'active_code': resource.active_code,
                'aliyun_url': resource.aliyun_url,
                'baidu_url': resource.baidu_url,
                'tianyi_url': resource.tianyi_url,
                'category': {
                    'id': resource.category.id,
                    'name': resource.category.name,
                    'slug': resource.category.slug,
                    'description': resource.category.description,
                    'status': resource.category.status,
                    'create_time': resource.category.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                    'update_time': resource.category.update_time.strftime('%Y-%m-%d %H:%M:%S')
                },
                'cover': site_url + '/uploads' + resource.cover.cover.url,
                'status': resource.status,
                'tags': [{
                    'id': tag.id,
                    'name': tag.name,
                    'slug': tag.slug,
                    'description': tag.description,
                    'status': tag.status,
                    'create_time': tag.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                    'update_time': tag.update_time.strftime('%Y-%m-%d %H:%M:%S')
                } for tag in resource.tags.all()],
                'create_time': resource.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                'update_time': resource.update_time.strftime('%Y-%m-%d%H:%M:%S')
            })
        return JsonResponse({'code': 200, 'msg': '获取成功', 'data': resource_list})
    else:
        return JsonResponse({'code': 400, 'msg': '请求方式错误'})


def admin_resource_add(request):
    """
    管理员添加资源
    :param request:
    :return:
    """
    token = request.headers.get('Authorization')
    payload, flag = check_token(token)
    if not flag:
        return JsonResponse({'code': 401, 'msg': 'token认证失败'})
    if not check_admin(token):
        return JsonResponse({'code': 401, 'msg': '权限不足'})
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        if not data:
            return JsonResponse({'code': 400, 'msg': '请填写完整内容'})
        if data.get('name'):
            name = data.get('name')
        else:
            return JsonResponse({'code': 400, 'msg': '请填写资源名称'})

        if name in Resource.objects.values_list('name', flat=True):
            return JsonResponse({'code': 400, 'msg': '资源名称重复'})

        if data.get('description'):
            description = data.get('description')
        else:
            description = ''

        if data.get('content'):
            content = data.get('content')
        else:
            content = ''

        if data.get('active_code'):
            active_code = data.get('active_code')
        else:
            active_code = ''

        if data.get('aliUrl'):
            aliyun_url = data.get('aliUrl')
        else:
            aliyun_url = ''
        if data.get('baiduUrl'):
            baidu_url = data.get('baiduUrl')
        else:
            baidu_url = ''

        if data.get('tianyiUrl'):
            tianyi_url = data.get('tianyiUrl')
        else:
            tianyi_url = ''

        if data.get('category'):
            category = data.get('category')
        else:
            category = 1

        if data.get('cover'):
            covers = data.get('cover')
        else:
            covers = []

        if data.get('tag'):
            tags = data.get('tag')
        else:
            tags = []

        if data.get('image'):
            images = data.get('image')
        else:
            images = []

        if data.get('video'):
            videos = data.get('video')
        else:
            videos = []

        if data.get('status'):
            status = data.get('status')
        else:
            status = True

        resource = Resource.objects.create(name=name, description=description, active_code=active_code, content=content,
                                           aliyun_url=aliyun_url, baidu_url=baidu_url, tianyi_url=tianyi_url,
                                           category=Category.objects.filter(id=category).first(),
                                           status=status)
        for cover in covers:
            resource.cover = Cover.objects.filter(id=cover).first()

        for tag in tags:
            tag_obj = Tag.objects.filter(id=tag).first()
            if tag_obj:
                resource.tags.add(tag_obj)

        for image in images:
            image_obj = Image.objects.filter(id=image).first()
            if image_obj:
                resource.images.add(image_obj)

        for video in videos:
            video_obj = Video.objects.filter(id=video).first()
            if video_obj:
                resource.videos.add(video_obj)

        return JsonResponse({'code': 200, 'msg': '添加成功'})
    else:
        return JsonResponse({'code': 400, 'msg': '请求方式错误'})


def admin_resource_edit(request):
    """
    管理员编辑资源
    :param request:
    :return:
    """
    token = request.headers.get('Authorization')
    payload, flag = check_token(token)
    if not flag:
        return JsonResponse({'code': 401, 'msg': 'token认证失败'})
    if not check_admin(token):
        return JsonResponse({'code': 401, 'msg': '权限不足'})
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        if data.get('id'):
            id = data.get('id')
        else:
            return JsonResponse({'code': 400, 'msg': '请填写资源id'})
        if not Resource.objects.filter(id=id).exists():
            return JsonResponse({'code': 400, 'msg': '资源不存在'})
        if data.get('name'):
            name = data.get('name')
        else:
            return JsonResponse({'code': 400, 'msg': '请填写资源名称'})
        if name in Resource.objects.exclude(id=id).values_list('name', flat=True):
            return JsonResponse({'code': 400, 'msg': '资源名称重复'})
        if data.get('description'):
            description = data.get('description')
        else:
            description = ''

        if data.get('content'):
            content = data.get('content')
        else:
            content = ''

        if data.get('active_code'):
            active_code = data.get('active_code')
        else:
            active_code = ''

        if data.get('aliUrl'):
            aliyun_url = data.get('aliUrl')
        else:
            aliyun_url = ''

        if data.get('baiduUrl'):
            baidu_url = data.get('baiduUrl')
        else:
            baidu_url = ''

        if data.get('tianyiUrl'):
            tianyi_url = data.get('tianyiUrl')
        else:
            tianyi_url = ''

        if data.get('category'):
            category = data.get('category')
        else:
            category = 1

        if data.get('cover'):
            covers = data.get('cover')
        else:
            covers = []

        if data.get('tag'):
            tags = data.get('tag')
        else:
            tags = []

        if data.get('image'):
            images = data.get('image')
        else:
            images = []

        if data.get('video'):
            videos = data.get('video')
        else:
            videos = []

        if data.get('status'):
            status = data.get('status')
        else:
            status = True

        resource = Resource.objects.filter(id=id).first()
        resource.name = name
        resource.description = description
        resource.content = content
        resource.active_code = active_code
        resource.aliyun_url = aliyun_url
        resource.baidu_url = baidu_url
        resource.tianyi_url = tianyi_url
        resource.category = Category.objects.filter(id=category).first()
        resource.status = status
        resource.save()
        if covers:
            for cover in covers:
                resource.cover = Cover.objects.filter(id=cover).first()
        else:
            resource.cover = Cover.objects.filter(id=1).first()

        resource.tags.clear()
        for tag in tags:
            tag_obj = Tag.objects.filter(id=tag).first()
            if tag_obj:
                resource.tags.add(tag_obj)

        resource.images.clear()
        for image in images:
            image_obj = Image.objects.filter(id=image).first()
            if image_obj:
                resource.images.add(image_obj)

        resource.videos.clear()
        for video in videos:
            video_obj = Video.objects.filter(id=video).first()
            if video_obj:
                resource.videos.add(video_obj)

        resource.save()
        return JsonResponse({'code': 200, 'msg': '编辑成功'})
    else:
        return JsonResponse({'code': 400, 'msg': '请求方式错误'})


def admin_resource_delete(request):
    """
    管理员删除资源
    :param request:
    :return:
    """
    token = request.headers.get('Authorization')
    payload, flag = check_token(token)
    if not flag:
        return JsonResponse({'code': 401, 'msg': 'token认证失败'})
    if not check_admin(token):
        return JsonResponse({'code': 401, 'msg': '权限不足'})
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        if data.get('resourceId'):
            id = data.get('resourceId')
        else:
            return JsonResponse({'code': 400, 'msg': '请填写资源id'})
        if Resource.objects.filter(id=id).exists():
            Resource.objects.filter(id=id).delete()
            return JsonResponse({'code': 200, 'msg': '删除成功'})
        else:
            return JsonResponse({'code': 400, 'msg': '资源不存在'})
    else:
        return JsonResponse({'code': 400, 'msg': '请求方式错误'})


def admin_resource_add_multiple(request):
    """
    管理员批量添加资源
    :param request:
    :return:
    """
    token = request.headers.get('Authorization')
    payload, flag = check_token(token)
    if not flag:
        return JsonResponse({'code': 401, 'msg': 'token认证失败'})
    if not check_admin(token):
        return JsonResponse({'code': 401, 'msg': '权限不足'})
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'code': 400, 'msg': '请上传文件'})
        if file.name.split('.')[-1] != 'xlsx':
            return JsonResponse({'code': 400, 'msg': '请上传xlsx文件'})
        data = openpyxl.load_workbook(file)
        sheet = data.active
        if sheet.max_row == 1:
            return JsonResponse({'code': 400, 'msg': '请上传完整文件'})
        upload_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            no = row[0]
            name = row[1]
            baidu = row[2]
            tianyi = row[3]
            ali = row[4]
            active_code = row[5]

            if not baidu:
                baidu = ''
            if not tianyi:
                tianyi = ''
            if not ali:
                ali = ''
            if not active_code:
                active_code = ''

            if Resource.objects.filter(name=name).exists():
                upload_list.append({
                    'no': no,
                    'percentage': 50,
                    'status': 0,
                    'name': name,
                    'msg': '资源已存在'
                })
                continue
            elif not name:
                upload_list.append({
                    'no': no,
                    'percentage': 50,
                    'status': 0,
                    'name': name,
                    'msg': '资源名称不能为空'
                })
                continue
            else:
                Resource.objects.create(
                    name=name,
                    baidu_url=baidu,
                    tianyi_url=tianyi,
                    aliyun_url=ali,
                    active_code=active_code,
                    category=Category.objects.filter(id=1).first(),
                    cover=Cover.objects.filter(id=1).first(),
                )
                upload_list.append({
                    'no': no,
                    'status': 1,
                    'percentage': 100,
                    'name': name,
                    'msg': '添加成功'
                })
        return JsonResponse({'code': 200, 'msg': '添加成功', 'data': upload_list})
    else:
        return JsonResponse({'code': 400, 'msg': '请求方式错误'})
